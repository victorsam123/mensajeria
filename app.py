import os
import re
import uuid
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from functools import wraps
from datetime import datetime
from urllib.parse import quote
from zoneinfo import ZoneInfo
from flask import (
    Flask, g, render_template, request, redirect,
    url_for, session, flash, send_file
)
from sqlalchemy import inspect, text, func, case
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

# ─────────────────────────── Configuración ───────────────────────────
BASE_DIR    = os.path.abspath(os.path.dirname(__file__))
UPLOAD_DIR  = os.path.join(BASE_DIR, "uploads")
EXPORT_DIR  = os.path.join(BASE_DIR, "exports")
DB_PATH     = os.path.join(BASE_DIR, "database", "mensajeria.db")
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin").strip() or "admin"
ADMIN_INITIAL_PASSWORD = os.getenv("ADMIN_PASSWORD", "cambiar-esta-clave")
ADMIN_FORCE_RESET = os.getenv("ADMIN_FORCE_RESET", "").strip() == "1"
ZONA_HORARIA_PARAGUAY = ZoneInfo("America/Asuncion")


def obtener_database_uri() -> str:
    database_url = os.getenv("DATABASE_URL", "").strip()
    if database_url:
        # Railway puede exponer postgres://; SQLAlchemy espera postgresql://
        if database_url.startswith("postgres://"):
            database_url = database_url.replace("postgres://", "postgresql+psycopg2://", 1)
        elif database_url.startswith("postgresql://"):
            database_url = database_url.replace("postgresql://", "postgresql+psycopg2://", 1)
        return database_url

    return f"sqlite:///{DB_PATH}"

ALLOWED_EXTENSIONS = {"xlsx", "xls"}

app = Flask(__name__)
SECRET_KEY = os.getenv("SECRET_KEY", "").strip()
# En produccion conviene configurar SECRET_KEY en Railway, pero la app no debe caerse si falta.
if SECRET_KEY:
    app.secret_key = SECRET_KEY
elif os.getenv("RAILWAY_ENVIRONMENT"):
    app.secret_key = os.urandom(32).hex()
else:
    app.secret_key = "mensajeria-masiva-dev-key"

app.config["SQLALCHEMY_DATABASE_URI"]        = obtener_database_uri()
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["UPLOAD_FOLDER"]                  = UPLOAD_DIR
app.config["MAX_CONTENT_LENGTH"]             = 16 * 1024 * 1024  # 16 MB
app.config["SQLALCHEMY_ENGINE_OPTIONS"]      = {
    "pool_pre_ping": True,
}

db = SQLAlchemy(app)

for d in [UPLOAD_DIR, EXPORT_DIR, os.path.dirname(DB_PATH)]:
    os.makedirs(d, exist_ok=True)


# ─────────────────────────── Modelo ──────────────────────────────────
class ContactoTemporal(db.Model):
    __tablename__ = "contactos_temporales"

    id          = db.Column(db.Integer, primary_key=True)
    sesion_id   = db.Column(db.String(36), nullable=False, index=True)
    nombre      = db.Column(db.String(200))
    apellido    = db.Column(db.String(200))
    nombre_madre = db.Column(db.String(200))
    apellido_madre = db.Column(db.String(200))
    documento_madre = db.Column(db.String(100))
    edad_anios  = db.Column(db.String(50))
    telefono    = db.Column(db.String(50))
    estado      = db.Column(db.String(20), default="valido")   # valido | duplicado | sin_telefono
    creado_en   = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self):
        return {
            "id":       self.id,
            "nombre":   self.nombre   or "",
            "apellido": self.apellido or "",
            "nombre_madre": self.nombre_madre or "",
            "apellido_madre": self.apellido_madre or "",
            "documento_madre": self.documento_madre or "",
            "edad_anios": self.edad_anios or "",
            "telefono": self.telefono or "",
            "estado":   self.estado,
        }


class Usuario(db.Model):
    __tablename__ = "usuarios"

    id            = db.Column(db.Integer, primary_key=True)
    username      = db.Column(db.String(80), nullable=False, unique=True, index=True)
    password_hash = db.Column(db.String(255), nullable=False)
    is_admin      = db.Column(db.Boolean, default=False)
    activo        = db.Column(db.Boolean, default=True)
    creado_en     = db.Column(db.DateTime, default=datetime.utcnow)

    def set_password(self, password: str) -> None:
        self.password_hash = generate_password_hash(password)

    def check_password(self, password: str) -> bool:
        return check_password_hash(self.password_hash, password)


class SesionUsuario(db.Model):
    __tablename__ = "sesiones_usuario"

    id                   = db.Column(db.Integer, primary_key=True)
    usuario_id           = db.Column(db.Integer, db.ForeignKey("usuarios.id"), nullable=False, index=True)
    username             = db.Column(db.String(80), nullable=False, index=True)
    fecha                = db.Column(db.Date, nullable=False, index=True)
    hora_inicio          = db.Column(db.DateTime, nullable=False)
    hora_fin             = db.Column(db.DateTime)
    total_procesados     = db.Column(db.Integer, default=0)
    total_validos        = db.Column(db.Integer, default=0)
    total_sin_telefono   = db.Column(db.Integer, default=0)
    total_duplicados     = db.Column(db.Integer, default=0)
    archivo_origen       = db.Column(db.String(255))
    hoja_origen          = db.Column(db.String(255))
    creado_en            = db.Column(db.DateTime, default=datetime.utcnow)


class RegistroEnvioUsuario(db.Model):
    __tablename__ = "registros_envio_usuario"

    id              = db.Column(db.Integer, primary_key=True)
    usuario_id      = db.Column(db.Integer, db.ForeignKey("usuarios.id"), nullable=False, index=True)
    username        = db.Column(db.String(80), nullable=False, index=True)
    sesion_id       = db.Column(db.Integer, db.ForeignKey("sesiones_usuario.id"), nullable=False, index=True)
    fecha           = db.Column(db.Date, nullable=False, index=True)
    hora            = db.Column(db.DateTime, nullable=False)
    nombre          = db.Column(db.String(200))
    apellido        = db.Column(db.String(200))
    nombre_madre    = db.Column(db.String(200))
    apellido_madre  = db.Column(db.String(200))
    documento_madre = db.Column(db.String(100))
    edad_anios      = db.Column(db.String(50))
    telefono        = db.Column(db.String(50))
    estado          = db.Column(db.String(20), nullable=False, index=True)
    archivo_origen  = db.Column(db.String(255), index=True)
    hoja_origen     = db.Column(db.String(255), index=True)
    creado_en       = db.Column(db.DateTime, default=datetime.utcnow)


class RegistroMensajeWhatsApp(db.Model):
    __tablename__ = "registros_mensaje_whatsapp"

    id                = db.Column(db.Integer, primary_key=True)
    usuario_id        = db.Column(db.Integer, db.ForeignKey("usuarios.id"), nullable=False, index=True)
    username          = db.Column(db.String(80), nullable=False, index=True)
    sesion_usuario_id = db.Column(db.Integer, db.ForeignKey("sesiones_usuario.id"), index=True)
    contacto_id       = db.Column(db.Integer, index=True)
    fecha             = db.Column(db.Date, nullable=False, index=True)
    hora              = db.Column(db.DateTime, nullable=False)
    nombre_madre      = db.Column(db.String(200))
    nombre_hijo       = db.Column(db.String(200))
    telefono          = db.Column(db.String(50), nullable=False)
    estado            = db.Column(db.String(30), nullable=False, default="Mensaje preparado")
    mensaje           = db.Column(db.Text, nullable=False)
    creado_en         = db.Column(db.DateTime, default=datetime.utcnow)


def ensure_admin_user() -> None:
    admin_exists = Usuario.query.filter_by(is_admin=True).first()
    if admin_exists and not ADMIN_FORCE_RESET:
        return

    admin = Usuario.query.filter_by(username=ADMIN_USERNAME).first()
    if not admin:
        admin = Usuario(username=ADMIN_USERNAME)
        db.session.add(admin)

    admin.is_admin = True
    admin.activo = True
    admin.set_password(ADMIN_INITIAL_PASSWORD)
    db.session.commit()


with app.app_context():
    db.create_all()
    inspector = inspect(db.engine)

    nombres_columnas_contactos = {
        c["name"]
        for c in inspector.get_columns("contactos_temporales")
    } if inspector.has_table("contactos_temporales") else set()
    if "documento_madre" not in nombres_columnas_contactos:
        db.session.execute(text("ALTER TABLE contactos_temporales ADD COLUMN documento_madre VARCHAR(100)"))
        db.session.commit()
    if "edad_anios" not in nombres_columnas_contactos:
        db.session.execute(text("ALTER TABLE contactos_temporales ADD COLUMN edad_anios VARCHAR(50)"))
        db.session.commit()
    if "nombre_madre" not in nombres_columnas_contactos:
        db.session.execute(text("ALTER TABLE contactos_temporales ADD COLUMN nombre_madre VARCHAR(200)"))
        db.session.commit()
    if "apellido_madre" not in nombres_columnas_contactos:
        db.session.execute(text("ALTER TABLE contactos_temporales ADD COLUMN apellido_madre VARCHAR(200)"))
        db.session.commit()

    nombres_columnas_registros = {
        c["name"]
        for c in inspector.get_columns("registros_envio_usuario")
    } if inspector.has_table("registros_envio_usuario") else set()
    if "nombre_madre" not in nombres_columnas_registros:
        db.session.execute(text("ALTER TABLE registros_envio_usuario ADD COLUMN nombre_madre VARCHAR(200)"))
        db.session.commit()
    if "apellido_madre" not in nombres_columnas_registros:
        db.session.execute(text("ALTER TABLE registros_envio_usuario ADD COLUMN apellido_madre VARCHAR(200)"))
        db.session.commit()

    nombres_columnas_mensajes = {
        c["name"]
        for c in inspector.get_columns("registros_mensaje_whatsapp")
    } if inspector.has_table("registros_mensaje_whatsapp") else set()
    if "nombre_hijo" not in nombres_columnas_mensajes:
        db.session.execute(text("ALTER TABLE registros_mensaje_whatsapp ADD COLUMN nombre_hijo VARCHAR(200)"))
        db.session.commit()

    nombres_columnas_registros_envio = {
        c["name"]
        for c in inspector.get_columns("registros_envio_usuario")
    } if inspector.has_table("registros_envio_usuario") else set()
    if "registros_envio_usuario" in inspector.get_table_names() and "edad_anios" not in nombres_columnas_registros_envio:
        db.session.execute(text("ALTER TABLE registros_envio_usuario ADD COLUMN edad_anios VARCHAR(50)"))
        db.session.commit()

    nombres_columnas_usuarios = {
        c["name"]
        for c in inspector.get_columns("usuarios")
    } if inspector.has_table("usuarios") else set()
    if "is_admin" not in nombres_columnas_usuarios:
        db.session.execute(text("ALTER TABLE usuarios ADD COLUMN is_admin BOOLEAN DEFAULT 0"))
        db.session.commit()

    ensure_admin_user()


# ─────────────────────────── Utilidades ──────────────────────────────
def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def obtener_motor_excel(ruta: str) -> str:
    extension = os.path.splitext(ruta)[1].lower()
    if extension == ".xls":
        return "xlrd"
    return "openpyxl"


def get_current_user() -> Usuario | None:
    user_id = session.get("user_id")
    if not user_id:
        return None

    user = db.session.get(Usuario, user_id)
    if not user or not user.activo:
        session.pop("user_id", None)
        session.pop("username", None)
        return None

    return user


def login_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if not get_current_user():
            flash("Debes iniciar sesión para continuar.", "warning")
            return redirect(url_for("login", next=request.path))
        return func(*args, **kwargs)

    return wrapper


def admin_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        user = get_current_user()
        if not user:
            flash("Debes iniciar sesión para continuar.", "warning")
            return redirect(url_for("login", next=request.path))
        if not user.is_admin:
            flash("Solo el administrador puede realizar esta acción.", "danger")
            return redirect(url_for("index"))
        return func(*args, **kwargs)

    return wrapper


def normalizar_next_url(next_url: str) -> str:
    if next_url and next_url.startswith("/") and not next_url.startswith("//"):
        return next_url
    return url_for("index")


def fecha_actual():
    return datetime.now().date()


def abrir_sesion_usuario(user: Usuario) -> SesionUsuario:
    ahora = datetime.now()
    sesiones_abiertas = SesionUsuario.query.filter_by(
        usuario_id=user.id,
        hora_fin=None,
    ).all()
    for sesion_abierta in sesiones_abiertas:
        sesion_abierta.hora_fin = ahora

    sesion_usuario = SesionUsuario(
        usuario_id=user.id,
        username=user.username,
        fecha=ahora.date(),
        hora_inicio=ahora,
    )
    db.session.add(sesion_usuario)
    db.session.commit()
    return sesion_usuario


def obtener_sesion_usuario_actual() -> SesionUsuario | None:
    sesion_usuario_id = session.get("sesion_usuario_id")
    if not sesion_usuario_id:
        return None
    return db.session.get(SesionUsuario, sesion_usuario_id)


def cerrar_sesion_usuario_actual() -> None:
    sesion_usuario = obtener_sesion_usuario_actual()
    if sesion_usuario and not sesion_usuario.hora_fin:
        sesion_usuario.hora_fin = datetime.now()
        db.session.commit()
    session.pop("sesion_usuario_id", None)


def actualizar_totales_sesion_usuario(sesion_usuario: SesionUsuario, stats: dict, archivo: str, hoja: str) -> None:
    sesion_usuario.total_procesados = (sesion_usuario.total_procesados or 0) + int(stats.get("total", 0))
    sesion_usuario.total_validos = (sesion_usuario.total_validos or 0) + int(stats.get("validos", 0))
    sesion_usuario.total_sin_telefono = (sesion_usuario.total_sin_telefono or 0) + int(stats.get("sin_tel", 0))
    sesion_usuario.total_duplicados = (sesion_usuario.total_duplicados or 0) + int(stats.get("duplicados", 0))
    sesion_usuario.archivo_origen = archivo
    sesion_usuario.hoja_origen = hoja


def resumen_registros(query):
    registros = query.all()
    total = len(registros)
    validos = sum(1 for r in registros if r.estado == "valido")
    sin_telefono = sum(1 for r in registros if r.estado == "sin_telefono")
    duplicados = sum(1 for r in registros if r.estado == "duplicado")
    return {
        "total": total,
        "validos": validos,
        "sin_telefono": sin_telefono,
        "duplicados": duplicados,
    }


def resumen_usuario_hoy(user: Usuario) -> dict:
    hoy = fecha_actual()
    return resumen_registros(
        RegistroEnvioUsuario.query.filter_by(usuario_id=user.id, fecha=hoy)
    )


def resumen_general_hoy() -> dict:
    hoy = fecha_actual()
    return resumen_registros(
        RegistroEnvioUsuario.query.filter_by(fecha=hoy)
    )


def totales_por_usuario_hoy():
    hoy = fecha_actual()
    filas = db.session.query(
        RegistroEnvioUsuario.username,
        func.count(RegistroEnvioUsuario.id),
        func.sum(case((RegistroEnvioUsuario.estado == "valido", 1), else_=0)),
    ).filter(
        RegistroEnvioUsuario.fecha == hoy
    ).group_by(
        RegistroEnvioUsuario.username
    ).order_by(
        RegistroEnvioUsuario.username.asc()
    ).all()
    return [
        {
            "username": username,
            "total": total or 0,
            "validos": validos or 0,
        }
        for username, total, validos in filas
    ]


def usuarios_activos_hoy() -> int:
    hoy = fecha_actual()
    return SesionUsuario.query.filter(
        SesionUsuario.fecha == hoy,
        SesionUsuario.hora_fin.is_(None),
    ).count()


def total_mensajes_preparados(fecha, user: Usuario | None = None) -> int:
    query = RegistroMensajeWhatsApp.query.filter_by(fecha=fecha)
    if user and not user.is_admin:
        query = query.filter_by(usuario_id=user.id)
    elif user:
        query = query.filter_by(usuario_id=user.id)
    return query.count()


def total_mensajes_preparados_hoy(user: Usuario | None = None) -> int:
    return total_mensajes_preparados(fecha_actual(), user)


def totales_mensajes_por_usuario(fecha, user: Usuario):
    query = db.session.query(
        RegistroMensajeWhatsApp.username,
        func.count(RegistroMensajeWhatsApp.id),
    ).filter(
        RegistroMensajeWhatsApp.fecha == fecha
    )
    if not user.is_admin:
        query = query.filter(RegistroMensajeWhatsApp.usuario_id == user.id)

    filas = query.group_by(
        RegistroMensajeWhatsApp.username
    ).order_by(
        RegistroMensajeWhatsApp.username.asc()
    ).all()

    return [
        {
            "username": username,
            "total": total or 0,
        }
        for username, total in filas
    ]


def saludo_institucional(ahora: datetime | None = None) -> str:
    ahora = ahora or datetime.now(ZONA_HORARIA_PARAGUAY)
    if ahora.tzinfo is not None:
        ahora = ahora.astimezone(ZONA_HORARIA_PARAGUAY)
    return "Buenos días" if ahora.hour < 12 else "Buenas tardes"


def unir_nombre_apellido(nombre: str, apellido: str = "") -> str:
    return " ".join(parte.strip() for parte in [nombre or "", apellido or ""] if parte and parte.strip())


def nombre_remitente(username: str = "") -> str:
    remitentes = {
        "victor": "Lic. Victor",
        "sonia": "Lic. Sonia",
    }
    return remitentes.get((username or "").strip().lower(), "")


def construir_mensaje_whatsapp(
    nombre_madre: str,
    nombre_hijo: str = "",
    ahora: datetime | None = None,
    username: str = "",
) -> str:
    nombre = (nombre_madre or "").strip() or "madre"
    referencia_hijo = f"su hijo/a {nombre_hijo.strip()}" if nombre_hijo and nombre_hijo.strip() else "su hijo/a"
    saludo = saludo_institucional(ahora)
    remitente = nombre_remitente(username)
    cierre = f"Muchas gracias.\n\nAtentamente,\n{remitente}" if remitente else "Muchas gracias."
    return (
        f"{saludo} Sr./Sra. {nombre}:\n\n"
        "Le saludo desde el Servicio de Vacunación del Hospital Regional de Ciudad del Este.\n\n"
        f"Nos comunicamos para recordarle que {referencia_hijo} registra vacunas pendientes que son necesarias "
        "para mantener su esquema de vacunación al día.\n\n"
        "Puede acercarse al Hospital Regional de Ciudad del Este de lunes a lunes, en el horario de 07:00 a 17:00 horas.\n\n"
        "Si lo prefiere, también puede enviarnos su ubicación para coordinar una visita domiciliaria.\n\n"
        "En caso de que su hijo/a ya cuente con todas las vacunas al día, favor omitir este mensaje.\n\n"
        f"{cierre}"
    )


def normalizar_telefono_whatsapp(telefono: str) -> str:
    return re.sub(r"\D", "", telefono or "")


@app.before_request
def load_current_user():
    g.current_user = get_current_user()


@app.context_processor
def inject_auth_context():
    return {
        "current_user": g.get("current_user"),
        "has_users": Usuario.query.count() > 0,
    }


# Palabras clave para detectar columnas automáticamente.
# Se prioriza coincidencia exacta (strip+lower) antes de coincidencia parcial.
KEYWORDS = {
    # Prioridad: columnas exactas del formato CVS → luego genéricas
    "nombre": [
        "madre nombre1", "nombre madre", "nombre de la madre",
        "madre nombre", "nombre1", "nombre", "nombres", "primer nombre",
    ],
    "apellido": [
        "madre apellido1", "apellido madre", "apellido de la madre",
        "madre apellido", "apellido1", "apellido", "apellidos", "primer apellido",
    ],
    "nombre_madre": [
        "madre_nombre1", "madre_nombre",
        "madre nombre1", "nombre madre", "nombre de la madre",
        "madre nombre", "nombres madre", "madre nombres",
    ],
    "apellido_madre": [
        "madre_apellido1", "madre_apellido",
        "madre apellido1", "apellido madre", "apellido de la madre",
        "madre apellido", "apellidos madre", "madre apellidos",
    ],
    "nombre_hijo": [
        "nombre1", "nombre", "nombres", "primer nombre",
        "persona nombre1", "persona nombre", "nombre persona",
        "hijo nombre", "nombre hijo", "nombre del hijo", "nombre de hijo",
        "hija nombre", "nombre hija", "nombre de la hija", "nombre de hija",
        "niño nombre", "nombre niño", "niña nombre", "nombre niña",
        "beneficiario nombre", "nombre beneficiario",
    ],
    "apellido_hijo": [
        "apellido1", "apellido", "apellidos", "primer apellido",
        "persona apellido1", "persona apellido", "apellido persona",
        "hijo apellido", "apellido hijo", "apellido del hijo", "apellido de hijo",
        "hija apellido", "apellido hija", "apellido de la hija", "apellido de hija",
        "niño apellido", "apellido niño", "niña apellido", "apellido niña",
        "beneficiario apellido", "apellido beneficiario",
    ],
    "documento_madre": [
        "madre_documento", "documento_madre", "madre documento", "documento madre", "doc madre",
        "documento de la madre", "madre doc", "cedula madre", "cédula madre",
    ],
    "edad_anios": [
        "edadaños", "edad años", "edad anios", "edad anos",
        "edad_anios", "edad_anos", "edad", "años", "anos",
    ],
    "telefono": [
        "telefono", "teléfono", "celular", "cel", "movil", "móvil",
        "numero", "número", "whatsapp", "contacto", "tel",
        "telefono madre", "teléfono madre", "celular madre",
        "numero de contacto", "número de contacto",
    ],
}


def detectar_columna(columnas: list[str], tipo: str) -> str | None:
    """
    Retorna la primera columna que coincida con las palabras clave del tipo dado.
    Busca primero por coincidencia exacta, luego por coincidencia parcial,
    respetando el orden de prioridad definido en KEYWORDS.
    """
    keywords = KEYWORDS.get(tipo, [])
    col_map = {c.lower().strip(): c for c in columnas}
    evitar_madre = tipo in {"nombre_hijo", "apellido_hijo"}

    # 1) Coincidencia exacta (prioridad máxima, respeta orden de keywords)
    for kw in keywords:
        if kw in col_map:
            return col_map[kw]

    # 2) Coincidencia parcial
    for kw in keywords:
        for col_lower, col_orig in col_map.items():
            if evitar_madre and "madre" in col_lower:
                continue
            if kw in col_lower or col_lower in kw:
                return col_orig

    return None


def normalizar_telefono(valor) -> str:
    """Limpia y normaliza un número telefónico."""
    if pd.isna(valor):
        return ""
    tel = str(valor).strip()
    # Quitar caracteres que no sean dígitos ni el + inicial
    tel = re.sub(r"[^\d+]", "", tel)
    # Si queda vacío o solo tiene + devolver cadena vacía
    if not tel or tel == "+":
        return ""
    # Mínimo 7 dígitos para ser un número válido (rechaza ej. solo "595")
    solo_digitos = re.sub(r"[^\d]", "", tel)
    if len(solo_digitos) < 7:
        return ""
    return tel


def limpiar_texto(valor) -> str:
    if pd.isna(valor):
        return ""
    return str(valor).strip()


def normalizar_edad_meses(valor) -> str:
    texto = limpiar_texto(valor)
    if not texto:
        return ""

    coincidencia = re.fullmatch(r"(\d+(?:[.,]\d+)?)\s*(?:mes|meses)?", texto.lower())
    if not coincidencia:
        return texto

    meses_totales = int(round(float(coincidencia.group(1).replace(",", "."))))
    anos, meses = divmod(meses_totales, 12)
    partes = []
    if anos:
        partes.append(f"{anos} año" if anos == 1 else f"{anos} años")
    if meses:
        partes.append(f"{meses} mes" if meses == 1 else f"{meses} meses")
    return " y ".join(partes) if partes else "0 meses"


def _detectar_fila_encabezados(ruta: str, hoja: str) -> int:
    """
    Lee una hoja detectando automáticamente la fila de encabezados.
    Soporta archivos donde la primera fila es un título (ej. formato CVS).
    """
    engine = obtener_motor_excel(ruta)

    # Leer sin encabezado para escanear las primeras filas
    preview = pd.read_excel(ruta, sheet_name=hoja, engine=engine,
                            header=None, nrows=10, dtype=str)

    header_row = 0  # valor por defecto
    for idx, row in preview.iterrows():
        # Buscar la fila donde la mayoría de celdas son texto no-vacío
        valores = [str(v).strip() for v in row if pd.notna(v) and str(v).strip()]
        if len(valores) >= 3:
            # Si al menos una celda contiene palabra clave de encabezado real
            row_lower = " ".join(v.lower() for v in valores)
            if any(kw in row_lower for kw in
                   ["nombre", "apellido", "telefono", "id persona", "documento", "celular", "edad"]):
                header_row = idx
                break

    return header_row


def _leer_hoja_detectando_header(ruta: str, hoja: str, nrows: int | None = None) -> pd.DataFrame:
    """
    Lee una hoja detectando automaticamente la fila de encabezados.
    Soporta archivos donde la primera fila es un titulo (ej. formato CVS).
    """
    engine = obtener_motor_excel(ruta)
    header_row = _detectar_fila_encabezados(ruta, hoja)

    df = pd.read_excel(ruta, sheet_name=hoja, engine=engine,
                       header=header_row, dtype=str, nrows=nrows)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how="all").reset_index(drop=True)
    return df


def obtener_hojas_activas(ruta: str) -> list[str]:
    """Devuelve solo hojas visibles/activas del archivo Excel."""
    if obtener_motor_excel(ruta) == "xlrd":
        xls = pd.ExcelFile(ruta, engine="xlrd")
        return xls.sheet_names

    wb = load_workbook(ruta, read_only=True, data_only=True)
    try:
        return [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
    finally:
        wb.close()


def contar_filas_hoja_excel(ruta: str, hoja: str, header_row: int) -> int:
    """Cuenta filas de datos sin cargar toda la hoja en memoria."""
    if obtener_motor_excel(ruta) == "xlrd":
        df = _leer_hoja_detectando_header(ruta, hoja)
        return len(df)

    try:
        wb = load_workbook(ruta, read_only=True, data_only=True)
    except Exception:
        return 0

    try:
        ws = wb[hoja]
        return max((ws.max_row or 0) - header_row - 1, 0)
    finally:
        wb.close()


def ordenar_valores_texto(valores: list[str]) -> list[str]:
    def clave(valor: str):
        texto = str(valor).strip()
        try:
            return (0, float(texto.replace(",", ".")))
        except ValueError:
            return (1, texto.lower())

    return sorted({str(v).strip() for v in valores if str(v).strip()}, key=clave)


def _normalizar_rango_filtro(ref: str) -> str:
    if not ref:
        return ""
    return ref.split("!")[-1].replace("$", "")


def obtener_columnas_con_filtro_excel(ruta: str, hoja: str, header_row: int, columnas: list[str]) -> set[str]:
    """Detecta columnas incluidas en autofiltros o tablas de la hoja cargada."""
    if obtener_motor_excel(ruta) == "xlrd":
        return set()

    try:
        wb = load_workbook(ruta, read_only=False, data_only=True)
    except Exception:
        return set()

    try:
        ws = wb[hoja]
        header_excel_row = header_row + 1
        columnas_filtradas = set()
        rangos = []

        auto_filter_ref = _normalizar_rango_filtro(getattr(ws.auto_filter, "ref", "") or "")
        if auto_filter_ref:
            rangos.append(auto_filter_ref)

        for table in ws.tables.values():
            table_ref = _normalizar_rango_filtro(getattr(table, "ref", "") or "")
            if table_ref:
                rangos.append(table_ref)

        for ref in rangos:
            try:
                min_col, min_row, max_col, _ = range_boundaries(ref)
            except ValueError:
                continue
            if min_row != header_excel_row:
                continue
            for col_index in range(min_col, max_col + 1):
                nombre = limpiar_texto(ws.cell(row=header_excel_row, column=col_index).value)
                if nombre in columnas:
                    columnas_filtradas.add(nombre)

        return columnas_filtradas
    finally:
        wb.close()


def obtener_filtros_hoja(df: pd.DataFrame, columnas_filtradas: set[str], limite: int = 200) -> list[dict]:
    filtros = []
    for columna in df.columns:
        if columna not in columnas_filtradas:
            continue
        valores = ordenar_valores_texto(
            df[columna].dropna().astype(str).map(str.strip).tolist()
        )
        if not valores:
            continue
        filtros.append({
            "columna": columna,
            "valores": valores[:limite],
            "total_valores": len(valores),
            "limitado": len(valores) > limite,
        })
    return filtros


def analizar_hojas_excel(ruta: str, hojas: list[str]) -> list[dict]:
    """Analiza hojas para la pantalla de seleccion sin reabrir el archivo por cada hoja."""
    if obtener_motor_excel(ruta) == "xlrd":
        return [analizar_hoja_excel(ruta, hoja) for hoja in hojas]

    try:
        wb = load_workbook(ruta, read_only=True, data_only=True)
    except Exception as exc:
        return [{
            "nombre": hoja,
            "columnas": [],
            "filas": 0,
            "col_nombre": "",
            "col_apellido": "",
            "col_nombre_madre": "",
            "col_apellido_madre": "",
            "col_doc_madre": "",
            "col_edad_anios": "",
            "col_telefono": "",
            "edad_valores": [],
            "filtros": [],
            "error": str(exc),
        } for hoja in hojas]

    try:
        hojas_info = []
        for hoja in hojas:
            info = {
                "nombre": hoja,
                "columnas": [],
                "filas": 0,
                "col_nombre": "",
                "col_apellido": "",
                "col_nombre_madre": "",
                "col_apellido_madre": "",
                "col_doc_madre": "",
                "col_edad_anios": "",
                "col_telefono": "",
                "edad_valores": [],
                "filtros": [],
                "error": "",
            }

            try:
                ws = wb[hoja]
                header_row = 1
                header_values = []
                for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
                    valores = [limpiar_texto(v) for v in row if limpiar_texto(v)]
                    if len(valores) >= 3:
                        row_lower = " ".join(v.lower() for v in valores)
                        if any(kw in row_lower for kw in [
                            "nombre", "apellido", "telefono", "id persona", "documento", "celular", "edad"
                        ]):
                            header_row = idx
                            header_values = [limpiar_texto(v) for v in row]
                            break

                if not header_values:
                    row = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), [])
                    header_values = [limpiar_texto(v) for v in row]

                columnas = [col for col in header_values if col]
                col_edad_anios = detectar_columna(columnas, "edad_anios") or ""
                col_nombre = detectar_columna(columnas, "nombre_hijo") or detectar_columna(columnas, "nombre") or ""
                col_apellido = detectar_columna(columnas, "apellido_hijo") or detectar_columna(columnas, "apellido") or ""

                info.update({
                    "columnas": columnas,
                    "filas": max((ws.max_row or 0) - header_row, 0),
                    "col_nombre": col_nombre,
                    "col_apellido": col_apellido,
                    "col_nombre_madre": detectar_columna(columnas, "nombre_madre") or "",
                    "col_apellido_madre": detectar_columna(columnas, "apellido_madre") or "",
                    "col_doc_madre": detectar_columna(columnas, "documento_madre") or "",
                    "col_edad_anios": col_edad_anios,
                    "col_telefono": detectar_columna(columnas, "telefono") or "",
                })

                if col_edad_anios and col_edad_anios in columnas:
                    edad_col_index = columnas.index(col_edad_anios) + 1
                    valores_edad = []
                    max_row = ws.max_row or header_row
                    for row in ws.iter_rows(
                        min_row=header_row + 1,
                        max_row=max_row,
                        min_col=edad_col_index,
                        max_col=edad_col_index,
                        values_only=True,
                    ):
                        valor = limpiar_texto(row[0] if row else "")
                        if valor:
                            valores_edad.append(normalizar_edad_meses(valor))
                    info["edad_valores"] = ordenar_valores_texto(valores_edad)
            except Exception as exc:
                info["error"] = str(exc)

            hojas_info.append(info)

        return hojas_info
    finally:
        wb.close()


def analizar_hoja_excel(ruta: str, hoja: str) -> dict:
    """Obtiene columnas y metadatos de una hoja antes de procesarla."""
    info = {
        "nombre": hoja,
        "columnas": [],
        "filas": 0,
        "col_nombre": "",
        "col_apellido": "",
        "col_nombre_madre": "",
        "col_apellido_madre": "",
        "col_doc_madre": "",
        "col_edad_anios": "",
        "col_telefono": "",
        "edad_valores": [],
        "filtros": [],
        "error": "",
    }

    try:
        header_row = _detectar_fila_encabezados(ruta, hoja)
        df = _leer_hoja_detectando_header(ruta, hoja)
    except Exception as exc:
        info["error"] = str(exc)
        return info

    columnas = list(df.columns)
    filas = contar_filas_hoja_excel(ruta, hoja, header_row)
    columnas_filtradas = set()
    if filas <= 5000:
        columnas_filtradas = obtener_columnas_con_filtro_excel(ruta, hoja, header_row, columnas)
    col_edad_anios = detectar_columna(columnas, "edad_anios") or ""
    col_nombre = detectar_columna(columnas, "nombre_hijo") or detectar_columna(columnas, "nombre") or ""
    col_apellido = detectar_columna(columnas, "apellido_hijo") or detectar_columna(columnas, "apellido") or ""
    info.update({
        "columnas": columnas,
        "filas": filas or len(df),
        "col_nombre": col_nombre,
        "col_apellido": col_apellido,
        "col_nombre_madre": detectar_columna(columnas, "nombre_madre") or "",
        "col_apellido_madre": detectar_columna(columnas, "apellido_madre") or "",
        "col_doc_madre": detectar_columna(columnas, "documento_madre") or "",
        "col_edad_anios": col_edad_anios,
        "col_telefono": detectar_columna(columnas, "telefono") or "",
        "filtros": obtener_filtros_hoja(df, columnas_filtradas),
    })

    if col_edad_anios:
        info["edad_valores"] = ordenar_valores_texto(
            df[col_edad_anios].dropna().map(normalizar_edad_meses).tolist()
        )

    return info


def procesar_dataframe(
    df: pd.DataFrame,
    col_nombre: str,
    col_apellido: str,
    col_telefono: str,
    col_doc_madre: str = "",
    col_edad_anios: str = "",
    col_nombre_madre: str = "",
    col_apellido_madre: str = "",
):
    """
    Filtra, limpia y devuelve una lista de dicts con estadísticas.
    """
    registros   = []
    total       = len(df)
    sin_tel     = 0
    duplicados  = 0
    telefonos_vistos = set()

    for _, row in df.iterrows():
        nombre   = limpiar_texto(row.get(col_nombre,   ""))
        apellido = limpiar_texto(row.get(col_apellido, ""))
        nombre_madre = limpiar_texto(row.get(col_nombre_madre, "")) if col_nombre_madre else ""
        apellido_madre = limpiar_texto(row.get(col_apellido_madre, "")) if col_apellido_madre else ""
        if not nombre_madre:
            nombre_madre = nombre
        if not apellido_madre:
            apellido_madre = apellido
        documento_madre = limpiar_texto(row.get(col_doc_madre, "")) if col_doc_madre else ""
        edad_anios = normalizar_edad_meses(row.get(col_edad_anios, "")) if col_edad_anios else ""
        telefono = normalizar_telefono(row.get(col_telefono, ""))

        if not telefono:
            sin_tel += 1
            registros.append({
                "nombre":   nombre,
                "apellido": apellido,
                "nombre_madre": nombre_madre,
                "apellido_madre": apellido_madre,
                "documento_madre": documento_madre,
                "edad_anios": edad_anios,
                "telefono": telefono,
                "estado":   "sin_telefono",
            })
            continue

        if telefono in telefonos_vistos:
            duplicados += 1
            registros.append({
                "nombre":   nombre,
                "apellido": apellido,
                "nombre_madre": nombre_madre,
                "apellido_madre": apellido_madre,
                "documento_madre": documento_madre,
                "edad_anios": edad_anios,
                "telefono": telefono,
                "estado":   "duplicado",
            })
            continue

        telefonos_vistos.add(telefono)
        registros.append({
            "nombre":   nombre,
            "apellido": apellido,
            "nombre_madre": nombre_madre,
            "apellido_madre": apellido_madre,
            "documento_madre": documento_madre,
            "edad_anios": edad_anios,
            "telefono": telefono,
            "estado":   "valido",
        })

    validos = total - sin_tel - duplicados
    stats = {
        "total":      total,
        "sin_tel":    sin_tel,
        "duplicados": duplicados,
        "validos":    validos,
    }
    return registros, stats


# ─────────────────────────── Rutas ───────────────────────────────────

@app.route("/")
@login_required
def index():
    current_user = get_current_user()
    return render_template(
        "index.html",
        resumen_hoy=resumen_usuario_hoy(current_user),
        resumen_general=resumen_general_hoy() if current_user.is_admin else None,
        totales_por_usuario=totales_por_usuario_hoy() if current_user.is_admin else [],
        usuarios_activos=usuarios_activos_hoy() if current_user.is_admin else 0,
        mensajes_enviados_hoy=total_mensajes_preparados_hoy(None if current_user.is_admin else current_user),
    )


@app.route("/login", methods=["GET", "POST"])
def login():
    if g.get("current_user"):
        return redirect(url_for("index"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        if not username or not password:
            flash("Ingresa usuario y contraseña.", "warning")
            return redirect(url_for("login"))

        user = Usuario.query.filter_by(username=username).first()
        if not user or not user.activo or not user.check_password(password):
            flash("Credenciales inválidas.", "danger")
            return redirect(url_for("login"))

        session["user_id"] = user.id
        session["username"] = user.username
        sesion_usuario = abrir_sesion_usuario(user)
        session["sesion_usuario_id"] = sesion_usuario.id
        next_url_raw = request.form.get("next") or request.args.get("next") or ""
        next_url = normalizar_next_url(next_url_raw)
        flash(f"Bienvenido, {user.username}.", "success")
        return redirect(next_url)

    return render_template("login.html", next_url=request.args.get("next", ""))


@app.route("/register", methods=["GET", "POST"])
@admin_required
def register():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        confirm  = request.form.get("confirm_password", "")

        if len(username) < 3:
            flash("El usuario debe tener al menos 3 caracteres.", "warning")
            return redirect(url_for("register"))
        if len(password) < 6:
            flash("La contraseña debe tener al menos 6 caracteres.", "warning")
            return redirect(url_for("register"))
        if password != confirm:
            flash("Las contraseñas no coinciden.", "warning")
            return redirect(url_for("register"))
        if Usuario.query.filter_by(username=username).first():
            flash("Ese usuario ya existe.", "danger")
            return redirect(url_for("register"))

        user = Usuario(username=username, is_admin=False)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        flash("Usuario creado correctamente.", "success")
        return redirect(url_for("users"))

    return render_template("register.html")


@app.route("/logout")
@login_required
def logout():
    cerrar_sesion_usuario_actual()
    session.pop("user_id", None)
    session.pop("username", None)
    flash("Sesión cerrada.", "info")
    return redirect(url_for("login"))


@app.route("/users", methods=["GET", "POST"])
@admin_required
def users():
    if request.method == "POST":
        action = request.form.get("action", "create")

        if action == "create":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            confirm  = request.form.get("confirm_password", "")

            if len(username) < 3:
                flash("El usuario debe tener al menos 3 caracteres.", "warning")
                return redirect(url_for("users"))
            if len(password) < 6:
                flash("La contraseña debe tener al menos 6 caracteres.", "warning")
                return redirect(url_for("users"))
            if password != confirm:
                flash("Las contraseñas no coinciden.", "warning")
                return redirect(url_for("users"))
            if Usuario.query.filter_by(username=username).first():
                flash("Ese usuario ya existe.", "danger")
                return redirect(url_for("users"))

            user = Usuario(username=username, is_admin=False)
            user.set_password(password)
            db.session.add(user)
            db.session.commit()
            flash("Nuevo usuario creado correctamente.", "success")
            return redirect(url_for("users"))

        user_id = request.form.get("user_id", "").strip()
        if not user_id.isdigit():
            flash("Usuario inválido.", "danger")
            return redirect(url_for("users"))

        target = db.session.get(Usuario, int(user_id))
        if not target:
            flash("Usuario no encontrado.", "danger")
            return redirect(url_for("users"))

        current_user = get_current_user()
        if current_user and target.id == current_user.id:
            flash("No puedes modificar tu propia cuenta de administrador.", "warning")
            return redirect(url_for("users"))

        if target.is_admin:
            flash("La cuenta administrador no se puede modificar desde aquí.", "warning")
            return redirect(url_for("users"))

        if action == "toggle":
            target.activo = not target.activo
            db.session.commit()
            estado = "habilitado" if target.activo else "deshabilitado"
            flash(f"Usuario {estado} correctamente.", "success")
            return redirect(url_for("users"))

        if action == "delete":
            db.session.delete(target)
            db.session.commit()
            flash("Usuario eliminado correctamente.", "success")
            return redirect(url_for("users"))

        if action == "update":
            username = request.form.get("username", "").strip()
            new_password = request.form.get("new_password", "")

            if len(username) < 3:
                flash("El usuario debe tener al menos 3 caracteres.", "warning")
                return redirect(url_for("users"))

            existe = Usuario.query.filter(Usuario.username == username, Usuario.id != target.id).first()
            if existe:
                flash("Ese usuario ya existe.", "danger")
                return redirect(url_for("users"))

            target.username = username
            if new_password:
                if len(new_password) < 6:
                    flash("La nueva contraseña debe tener al menos 6 caracteres.", "warning")
                    return redirect(url_for("users"))
                target.set_password(new_password)

            db.session.commit()
            flash("Usuario actualizado correctamente.", "success")
            return redirect(url_for("users"))

        flash("Acción no soportada.", "danger")
        return redirect(url_for("users"))

    usuarios = Usuario.query.order_by(Usuario.creado_en.desc()).all()
    return render_template(
        "users.html",
        usuarios=usuarios,
        total_usuarios=len(usuarios),
    )


@app.route("/reporte-diario")
@login_required
def reporte_diario():
    fecha_raw = request.args.get("fecha", "").strip()
    try:
        fecha_reporte = datetime.strptime(fecha_raw, "%Y-%m-%d").date() if fecha_raw else fecha_actual()
    except ValueError:
        fecha_reporte = fecha_actual()

    current_user = get_current_user()
    filas_reporte = totales_mensajes_por_usuario(fecha_reporte, current_user)
    total_general = sum(fila["total"] for fila in filas_reporte)

    return render_template(
        "reporte_diario.html",
        fecha_reporte=fecha_reporte,
        filtros={
            "fecha": fecha_reporte.strftime("%Y-%m-%d"),
        },
        filas=filas_reporte,
        total_general=total_general,
    )


def limpiar_archivo_actual() -> None:
    """Limpia solo los datos temporales del archivo cargado actualmente."""
    sesion_id = session.get("sesion_id")
    if sesion_id:
        ContactoTemporal.query.filter_by(sesion_id=sesion_id).delete()
        db.session.commit()

    ruta = session.get("ruta")
    if ruta:
        try:
            ruta_abs = os.path.abspath(ruta)
            uploads_abs = os.path.abspath(app.config["UPLOAD_FOLDER"])
            if os.path.commonpath([ruta_abs, uploads_abs]) == uploads_abs and os.path.exists(ruta_abs):
                os.remove(ruta_abs)
        except (OSError, ValueError):
            pass

    claves_temporales = (
        "sesion_id",
        "archivo",
        "ruta",
        "hojas",
        "hoja",
        "stats",
        "columnas",
        "col_nombre",
        "col_apellido",
        "col_nombre_madre",
        "col_apellido_madre",
        "col_doc_madre",
        "col_edad_anios",
        "col_telefono",
        "edad_valores",
        "filtros_seleccionados",
    )
    for clave in claves_temporales:
        session.pop(clave, None)


@app.route("/upload", methods=["POST"])
@login_required
def upload():
    if "excel_file" not in request.files:
        flash("No se seleccionó ningún archivo.", "danger")
        return redirect(url_for("index"))

    archivo = request.files["excel_file"]

    if archivo.filename == "":
        flash("El archivo no tiene nombre.", "danger")
        return redirect(url_for("index"))

    if not allowed_file(archivo.filename):
        flash("Solo se permiten archivos .xlsx o .xls.", "danger")
        return redirect(url_for("index"))

    if session.get("sesion_id") or session.get("archivo"):
        limpiar_archivo_actual()

    nombre_seguro = secure_filename(archivo.filename)
    ruta          = os.path.join(app.config["UPLOAD_FOLDER"], nombre_seguro)
    archivo.save(ruta)

    # Leer nombres de hojas visibles/activas
    try:
        hojas = obtener_hojas_activas(ruta)
    except Exception as e:
        flash(f"Error al leer el archivo: {e}", "danger")
        return redirect(url_for("index"))

    if not hojas:
        flash("El archivo no contiene hojas.", "danger")
        return redirect(url_for("index"))

    session["archivo"]    = nombre_seguro
    session["ruta"]       = ruta
    session["hojas"]      = hojas
    flash("Nuevo archivo cargado correctamente.", "success")

    return redirect(url_for("select_sheet"))


@app.route("/select-sheet")
@login_required
def select_sheet():
    if "hojas" not in session:
        flash("Primero debes cargar un archivo.", "warning")
        return redirect(url_for("index"))

    ruta = session.get("ruta")
    if not ruta or not os.path.exists(ruta):
        flash("El archivo ya no estÃ¡ disponible. CÃ¡rgalo nuevamente.", "danger")
        return redirect(url_for("index"))

    hojas_info = analizar_hojas_excel(ruta, session.get("hojas", []))

    return render_template(
        "select_sheet.html",
        archivo=session.get("archivo"),
        hojas=session.get("hojas"),
        hojas_info=hojas_info,
        first_info=hojas_info[0] if hojas_info else None,
    )


@app.route("/process", methods=["POST"])
@login_required
def process():
    hoja = request.form.get("hoja", "").strip()
    if not hoja:
        flash("Debes seleccionar una hoja.", "warning")
        return redirect(url_for("select_sheet"))

    ruta = session.get("ruta")
    if not ruta or not os.path.exists(ruta):
        flash("El archivo ya no está disponible. Cárgalo nuevamente.", "danger")
        return redirect(url_for("index"))

    try:
        df = _leer_hoja_detectando_header(ruta, hoja)
    except Exception as e:
        flash(f"Error al leer la hoja '{hoja}': {e}", "danger")
        return redirect(url_for("select_sheet"))

    if df.empty:
        flash("La hoja seleccionada está vacía.", "warning")
        return redirect(url_for("select_sheet"))

    columnas = list(df.columns)

    # Detección automática de columnas
    col_nombre   = detectar_columna(columnas, "nombre")
    col_apellido = detectar_columna(columnas, "apellido")
    col_nombre_hijo = detectar_columna(columnas, "nombre_hijo")
    col_apellido_hijo = detectar_columna(columnas, "apellido_hijo")
    col_nombre_madre = detectar_columna(columnas, "nombre_madre")
    col_apellido_madre = detectar_columna(columnas, "apellido_madre")
    col_doc_madre = detectar_columna(columnas, "documento_madre")
    col_edad_anios = detectar_columna(columnas, "edad_anios")
    col_telefono = detectar_columna(columnas, "telefono")
    col_nombre = col_nombre_hijo or col_nombre
    col_apellido = col_apellido_hijo or col_apellido

    # Columnas enviadas manualmente (si el usuario las ajusta)
    col_nombre   = request.form.get("col_nombre",   col_nombre   or "")
    col_apellido = request.form.get("col_apellido", col_apellido or "")
    col_nombre_madre = request.form.get("col_nombre_madre", col_nombre_madre or "")
    col_apellido_madre = request.form.get("col_apellido_madre", col_apellido_madre or "")
    col_doc_madre = request.form.get("col_doc_madre", col_doc_madre or "")
    col_edad_anios = request.form.get("col_edad_anios", col_edad_anios or "")
    col_telefono = request.form.get("col_telefono", col_telefono or "")
    edad_valores = [
        valor.strip()
        for valor in request.form.getlist("edad_valores")
        if valor.strip()
    ]
    filtros_columnas = request.form.getlist("filtro_columnas")
    filtros_seleccionados = {}
    for index, filtro_columna in enumerate(filtros_columnas):
        filtro_columna = filtro_columna.strip()
        valores = [
            valor.strip()
            for valor in request.form.getlist(f"filtro_valores_{index}")
            if valor.strip()
        ]
        if filtro_columna and valores:
            filtros_seleccionados[filtro_columna] = valores

    columnas_set = set(columnas)
    col_nombre = col_nombre if col_nombre in columnas_set else ""
    col_apellido = col_apellido if col_apellido in columnas_set else ""
    col_nombre_madre = col_nombre_madre if col_nombre_madre in columnas_set else ""
    col_apellido_madre = col_apellido_madre if col_apellido_madre in columnas_set else ""
    col_doc_madre = col_doc_madre if col_doc_madre in columnas_set else ""
    col_edad_anios = col_edad_anios if col_edad_anios in columnas_set else ""
    col_telefono = col_telefono if col_telefono in columnas_set else ""

    # Si faltan columnas, redirigir con columnas detectadas para que el usuario ajuste
    if not col_telefono:
        flash(
            "No se detectó la columna de teléfono automáticamente. "
            "Selecciónala manualmente.",
            "warning",
        )
        session["hoja"]      = hoja
        session["columnas"]  = columnas
        session["col_nombre"]   = col_nombre
        session["col_apellido"] = col_apellido
        session["col_nombre_madre"] = col_nombre_madre
        session["col_apellido_madre"] = col_apellido_madre
        session["col_doc_madre"] = col_doc_madre
        session["col_edad_anios"] = col_edad_anios
        session["col_telefono"] = col_telefono
        session["edad_valores"] = edad_valores
        return redirect(url_for("map_columns"))

    if col_edad_anios and edad_valores:
        df = df[
            df[col_edad_anios].map(normalizar_edad_meses).isin(edad_valores)
        ].reset_index(drop=True)
        if df.empty:
            flash("No hay filas para procesar con el filtro de edad seleccionado.", "warning")
            return redirect(url_for("select_sheet"))

    for filtro_columna, valores in filtros_seleccionados.items():
        if filtro_columna not in columnas_set:
            continue
        df = df[
            df[filtro_columna].fillna("").astype(str).map(limpiar_texto).isin(valores)
        ].reset_index(drop=True)
        if df.empty:
            flash("No hay filas para procesar con los filtros seleccionados.", "warning")
            return redirect(url_for("select_sheet"))

    # Procesar
    registros, stats = procesar_dataframe(
        df,
        col_nombre,
        col_apellido,
        col_telefono,
        col_doc_madre,
        col_edad_anios,
        col_nombre_madre,
        col_apellido_madre,
    )

    # Guardar en BD con sesión única
    sesion_id = str(uuid.uuid4())
    session["sesion_id"] = sesion_id
    session["hoja"]      = hoja
    session["stats"]     = stats
    session["edad_valores"] = edad_valores
    session["filtros_seleccionados"] = filtros_seleccionados
    current_user = get_current_user()
    sesion_usuario = obtener_sesion_usuario_actual()
    if not sesion_usuario and current_user:
        sesion_usuario = abrir_sesion_usuario(current_user)
        session["sesion_usuario_id"] = sesion_usuario.id

    # Borrar registros anteriores de esta sesión (por si recargó)
    ContactoTemporal.query.filter_by(sesion_id=sesion_id).delete()

    archivo_origen = session.get("archivo", "")
    hora_registro = datetime.now()
    for r in registros:
        db.session.add(ContactoTemporal(
            sesion_id=sesion_id,
            nombre=r["nombre"],
            apellido=r["apellido"],
            nombre_madre=r["nombre_madre"],
            apellido_madre=r["apellido_madre"],
            documento_madre=r["documento_madre"],
            edad_anios=r["edad_anios"],
            telefono=r["telefono"],
            estado=r["estado"],
        ))
        if current_user and sesion_usuario:
            db.session.add(RegistroEnvioUsuario(
                usuario_id=current_user.id,
                username=current_user.username,
                sesion_id=sesion_usuario.id,
                fecha=hora_registro.date(),
                hora=hora_registro,
                nombre=r["nombre"],
                apellido=r["apellido"],
                nombre_madre=r["nombre_madre"],
                apellido_madre=r["apellido_madre"],
                documento_madre=r["documento_madre"],
                edad_anios=r["edad_anios"],
                telefono=r["telefono"],
                estado=r["estado"],
                archivo_origen=archivo_origen,
                hoja_origen=hoja,
            ))
    if sesion_usuario:
        actualizar_totales_sesion_usuario(sesion_usuario, stats, archivo_origen, hoja)
    db.session.commit()

    return redirect(url_for("preview"))


@app.route("/map-columns", methods=["GET", "POST"])
@login_required
def map_columns():
    """Permite al usuario seleccionar manualmente las columnas."""
    if "hojas" not in session:
        return redirect(url_for("index"))

    if request.method == "POST":
        return redirect(url_for("process"), code=307)  # reenviar con POST

    return render_template(
        "map_columns.html",
        archivo=session.get("archivo"),
        hoja=session.get("hoja"),
        columnas=session.get("columnas", []),
        col_nombre=session.get("col_nombre", ""),
        col_apellido=session.get("col_apellido", ""),
        col_nombre_madre=session.get("col_nombre_madre", ""),
        col_apellido_madre=session.get("col_apellido_madre", ""),
        col_doc_madre=session.get("col_doc_madre", ""),
        col_edad_anios=session.get("col_edad_anios", ""),
        col_telefono=session.get("col_telefono", ""),
    )


@app.route("/preview")
@login_required
def preview():
    sesion_id = session.get("sesion_id")
    if not sesion_id:
        flash("No hay datos procesados. Carga un archivo primero.", "warning")
        return redirect(url_for("index"))

    contactos = ContactoTemporal.query.filter_by(sesion_id=sesion_id).all()
    stats     = session.get("stats", {})

    return render_template(
        "preview.html",
        contactos=contactos,
        stats=stats,
        archivo=session.get("archivo"),
        hoja=session.get("hoja"),
    )


@app.route("/whatsapp/<int:contacto_id>")
@login_required
def preparar_whatsapp(contacto_id: int):
    sesion_id = session.get("sesion_id")
    contacto = db.session.get(ContactoTemporal, contacto_id)
    if not contacto or contacto.sesion_id != sesion_id:
        flash("Contacto no disponible para esta sesión.", "danger")
        return redirect(url_for("preview") if sesion_id else url_for("index"))

    if contacto.estado != "valido" or not contacto.telefono:
        flash("Solo se puede preparar WhatsApp para contactos válidos con teléfono.", "warning")
        return redirect(url_for("preview"))

    telefono = normalizar_telefono_whatsapp(contacto.telefono)
    if not telefono:
        flash("El teléfono del contacto no es válido para WhatsApp.", "warning")
        return redirect(url_for("preview"))

    current_user = get_current_user()
    ahora = datetime.now(ZONA_HORARIA_PARAGUAY)
    nombre_madre = unir_nombre_apellido(contacto.nombre_madre, contacto.apellido_madre)
    if not nombre_madre:
        nombre_madre = unir_nombre_apellido(contacto.nombre, contacto.apellido)
    nombre_hijo = unir_nombre_apellido(contacto.nombre, contacto.apellido)
    mensaje = construir_mensaje_whatsapp(nombre_madre, nombre_hijo, ahora, current_user.username)
    sesion_usuario = obtener_sesion_usuario_actual()

    db.session.add(RegistroMensajeWhatsApp(
        usuario_id=current_user.id,
        username=current_user.username,
        sesion_usuario_id=sesion_usuario.id if sesion_usuario else None,
        contacto_id=contacto.id,
        fecha=ahora.date(),
        hora=ahora,
        nombre_madre=nombre_madre,
        nombre_hijo=nombre_hijo,
        telefono=telefono,
        estado="Mensaje preparado",
        mensaje=mensaje,
    ))
    db.session.commit()

    mensaje_codificado = quote(mensaje)
    return redirect(f"https://wa.me/{telefono}?text={mensaje_codificado}")


@app.route("/export-respuestas/<fmt>")
@login_required
def export_respuestas(fmt: str):
    sesion_id = session.get("sesion_id")
    if not sesion_id:
        flash("No hay datos para generar la planilla de respuestas.", "warning")
        return redirect(url_for("index"))

    contactos = ContactoTemporal.query.filter_by(
        sesion_id=sesion_id, estado="valido"
    ).all()

    if not contactos:
        flash("No hay contactos validos para generar la planilla de respuestas.", "warning")
        return redirect(url_for("preview"))

    current_user = get_current_user()
    ahora = datetime.now(ZONA_HORARIA_PARAGUAY)
    filas = []
    for contacto in contactos:
        telefono = normalizar_telefono_whatsapp(contacto.telefono)
        nombre_madre = unir_nombre_apellido(contacto.nombre_madre, contacto.apellido_madre)
        if not nombre_madre:
            nombre_madre = unir_nombre_apellido(contacto.nombre, contacto.apellido)
        nombre_hijo = unir_nombre_apellido(contacto.nombre, contacto.apellido)
        mensaje = construir_mensaje_whatsapp(nombre_madre, nombre_hijo, ahora, current_user.username)

        filas.append({
            "nombre_hijo": contacto.nombre or "",
            "apellido_hijo": contacto.apellido or "",
            "nombre_madre": contacto.nombre_madre or "",
            "apellido_madre": contacto.apellido_madre or "",
            "documento_madre": contacto.documento_madre or "",
            "edad_anios": contacto.edad_anios or "",
            "telefono_original": contacto.telefono or "",
            "telefono_whatsapp": telefono,
            "enlace_whatsapp": f"https://wa.me/{telefono}?text={quote(mensaje)}" if telefono else "",
            "estado_respuesta": "Pendiente",
            "respuesta_recibida": "",
            "fecha_respuesta": "",
            "observaciones": "",
            "ubicacion_para_visita": "",
            "mensaje_preparado": mensaje,
        })

    df = pd.DataFrame(filas, columns=[
        "nombre_hijo", "apellido_hijo", "nombre_madre", "apellido_madre",
        "documento_madre", "edad_anios", "telefono_original", "telefono_whatsapp",
        "enlace_whatsapp", "estado_respuesta", "respuesta_recibida",
        "fecha_respuesta", "observaciones", "ubicacion_para_visita",
        "mensaje_preparado",
    ])

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_base = f"planilla_respuestas_{timestamp}"

    if fmt == "excel":
        ruta_export = os.path.join(EXPORT_DIR, f"{nombre_base}.xlsx")
        df.to_excel(ruta_export, index=False, engine="openpyxl")
        return send_file(
            ruta_export,
            as_attachment=True,
            download_name=f"{nombre_base}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    elif fmt == "csv":
        ruta_export = os.path.join(EXPORT_DIR, f"{nombre_base}.csv")
        df.to_csv(ruta_export, index=False, encoding="utf-8-sig")
        return send_file(
            ruta_export,
            as_attachment=True,
            download_name=f"{nombre_base}.csv",
            mimetype="text/csv",
        )

    flash("Formato de exportacion no soportado.", "danger")
    return redirect(url_for("preview"))


@app.route("/export/<fmt>")
@login_required
def export(fmt: str):
    sesion_id = session.get("sesion_id")
    if not sesion_id:
        flash("No hay datos para exportar.", "warning")
        return redirect(url_for("index"))

    # Solo exportar válidos
    contactos = ContactoTemporal.query.filter_by(
        sesion_id=sesion_id, estado="valido"
    ).all()

    if not contactos:
        flash("No hay contactos válidos para exportar.", "warning")
        return redirect(url_for("preview"))

    datos = [c.to_dict() for c in contactos]
    df    = pd.DataFrame(
        datos,
        columns=[
            "nombre", "apellido", "nombre_madre", "apellido_madre",
            "documento_madre", "edad_anios", "telefono", "estado",
        ],
    )
    df    = df.drop(columns=["id"], errors="ignore")

    timestamp    = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_base  = f"contactos_{timestamp}"

    if fmt == "excel":
        ruta_export = os.path.join(EXPORT_DIR, f"{nombre_base}.xlsx")
        df.to_excel(ruta_export, index=False, engine="openpyxl")
        return send_file(
            ruta_export,
            as_attachment=True,
            download_name=f"{nombre_base}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    elif fmt == "csv":
        ruta_export = os.path.join(EXPORT_DIR, f"{nombre_base}.csv")
        df.to_csv(ruta_export, index=False, encoding="utf-8-sig")
        return send_file(
            ruta_export,
            as_attachment=True,
            download_name=f"{nombre_base}.csv",
            mimetype="text/csv",
        )
    else:
        flash("Formato de exportación no soportado.", "danger")
        return redirect(url_for("preview"))


@app.route("/nuevo-archivo", methods=["POST"])
@login_required
def nuevo_archivo():
    limpiar_archivo_actual()
    flash("Archivo cerrado correctamente. Ya puede cargar un nuevo archivo.", "success")
    return redirect(url_for("index"))


@app.route("/cerrar-archivo", methods=["POST"])
@login_required
def cerrar_archivo():
    limpiar_archivo_actual()
    flash("Archivo cerrado correctamente. Ya puede cargar un nuevo archivo.", "success")
    return redirect(url_for("index"))


@app.route("/reset")
@login_required
def reset():
    limpiar_archivo_actual()
    flash("Archivo cerrado correctamente. Ya puede cargar un nuevo archivo.", "success")
    return redirect(url_for("index"))


if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=int(os.getenv("PORT", "5050")),
        debug=os.getenv("FLASK_DEBUG", "0") == "1",
    )
